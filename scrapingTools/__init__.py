from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import csv
import openpyxl
import time


def salva_em_csv(data, nome):
    # Extrair chaves dos dicionários
    cabecalho = ['label', 'link', 'phone_number', 'tipo_de_negocio', 'endereço']
    with open(nome+'.csv', 'w', newline='') as csvfile:
        escritor = csv.writer(csvfile, delimiter=',')

        # Escreva o cabeçalho
        escritor.writerow(cabecalho)

        # Escreva os dados
        for item in data:
            escritor.writerow([item[chave] for chave in cabecalho])
    return


def salva_em_excel(data, nome):
    # Extrair chaves dos dicionários
    cabecalho = ['label', 'link', 'phone_number', 'tipo_de_negocio', 'endereço']

    # Criar pasta de trabalho e planilha
    wb = openpyxl.Workbook()
    ws = wb.active

    # Escrever o cabeçalho
    for i, coluna in enumerate(cabecalho):
        ws.cell(row=1, column=i+1, value=coluna)

    # Escrever os dados
    for i, item in enumerate(data, start=2):
        for j, chave in enumerate(cabecalho):
            ws.cell(row=i, column=j+1, value=item[chave])

    # Salvar a pasta de trabalho
    wb.save(nome+'.xlsx')


def devolve_links(html):
    soup = BeautifulSoup(html, "html.parser")
    # Encontrar todos os elementos `a` com classe "hfpxzc"
    try:
        elementos_a = soup.find_all("div", class_="Nv2PK")
    except:
        return None

    lista=[]
    if elementos_a:
        print('elementos_a ',len(elementos_a))
        for elemento_a in elementos_a:
            infos={}
            label_link = elemento_a.find('a', class_='hfpxzc')
            infos['label'] = label_link['aria-label']
            infos['link'] = label_link['href']
            try:
                infos['phone_number'] = elemento_a.find('span', class_='UsdlK').text
            except:
                infos['phone_number'] = None
            divs = elemento_a.find_all('div', class_='W4Efsd')
            # Extrair o texto das divs
            for nn, div in enumerate(divs):
                for n,span in enumerate(div.find_all('span')):
                    if nn==2:
                        if n==0:
                            infos['tipo_de_negocio']=span.get_text(strip=True)
                        if n==4:
                            infos['endereço']=span.get_text(strip=True)
            lista.append(infos)
    return lista


def scrool_infinito(driver,deep,min_list_elements, max_repetition,atual=0,lista=[]):
    print(deep,min_list_elements, max_repetition,atual,len(lista))
    element = driver.find_elements(By.CLASS_NAME, 'm6QErb') #scrool
    for nn,i in enumerate(element):
        try:
            i.send_keys(Keys.PAGE_DOWN)
            time.sleep(.5)
            html = driver.page_source  # Capture HTML after new content loads
            for i in devolve_links(html):
                if i not in lista:
                    lista.append(i)
        except:
            pass
    if min_list_elements<len(lista):
        return lista
    if max_repetition<atual:
        return lista
    return scrool_infinito(driver,deep,min_list_elements, max_repetition,atual+1,lista)
    # Criar a sopa BeautifulSoup


def pesquisa_inicial(pesquisa, deep, silent=True,num_min_registros=20,max_try=200):
    # Criar um objeto do webdriver
    print(pesquisa)
    chrome_options = Options()
    if silent:
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")

    # Criar um objeto do webdriver com as opções configuradas
    driver = webdriver.Chrome(options=chrome_options)
    #driver = webdriver.Chrome()
    # URL base do Google Maps
    url_base = "https://www.google.com/maps/search/"
    # Abrir o navegador
    driver.get(url_base + pesquisa)
    # Esperar até que o elemento de classe específica "Nv2PK" seja visível na página
    try:
        WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.CLASS_NAME, "Nv2PK")))
    except:
        return None
    
    lista = scrool_infinito(driver,deep,num_min_registros, max_try,lista=[]) #adicionar 
    # Fechar o navegador
    driver.quit()
    return lista